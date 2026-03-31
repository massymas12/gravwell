from __future__ import annotations

import csv
import io
import ipaddress
import re
from pathlib import Path

from gravwell.models.dataclasses import ParseResult
from gravwell.parsers.base import BaseParser

# ---------------------------------------------------------------------------
# Column name aliases
# ---------------------------------------------------------------------------
# phpIPAM exports "Subnet" and "Description" exactly; also accept generic names.
_CIDR_ALIASES: set[str] = {
    "subnet", "network", "cidr", "prefix",
    "subnet_address", "network_address", "ip_subnet",
}
_LABEL_ALIASES: set[str] = {
    "description", "name", "label", "subnet_name",
    "subnet_description", "comment", "remarks",
}

# phpIPAM prepends ">" hierarchy markers to child subnet descriptions;
# strip any leading ">" chars and whitespace.
_HIERARCHY_PREFIX = re.compile(r"^[>\s]+")


def _normalise_header(h: str) -> str:
    return h.strip().lower()


def _find_col(norm_headers: list[str], aliases: set[str]) -> str | None:
    for h in norm_headers:
        if h in aliases:
            return h
    return None


def _is_valid_cidr(value: str) -> bool:
    try:
        ipaddress.ip_network(value.strip(), strict=False)
        return True
    except ValueError:
        return False


def _clean_label(raw: str) -> str:
    return _HIERARCHY_PREFIX.sub("", raw).strip()


def _canonical_cidr(cidr: str) -> str:
    return str(ipaddress.ip_network(cidr.strip(), strict=False))


# ---------------------------------------------------------------------------
# Shared parsing logic (works on an iterable of {norm_header: value} dicts)
# ---------------------------------------------------------------------------

def _parse_rows(
    rows: list[dict[str, str]],
    norm_headers: list[str],
    result: ParseResult,
) -> None:
    cidr_col = _find_col(norm_headers, _CIDR_ALIASES)
    label_col = _find_col(norm_headers, _LABEL_ALIASES)
    if not cidr_col or not label_col:
        result.errors.append(
            "Could not find required columns. "
            f"Expected a subnet column ({', '.join(sorted(_CIDR_ALIASES))}) "
            f"and a label column ({', '.join(sorted(_LABEL_ALIASES))})."
        )
        return

    for row in rows:
        cidr_raw = (row.get(cidr_col) or "").strip()
        label_raw = (row.get(label_col) or "").strip()
        if not cidr_raw or not _is_valid_cidr(cidr_raw):
            continue
        canonical = _canonical_cidr(cidr_raw)
        label = _clean_label(label_raw)
        # Prefer existing entry (first row wins) unless new label is richer
        existing = result.subnet_labels.get(canonical, "")
        if not existing or (label and len(label) > len(existing)):
            result.subnet_labels[canonical] = label or canonical


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

class IPAMParser(BaseParser):
    name = "ipam"

    @classmethod
    def can_parse(cls, filepath: Path) -> bool:
        suffix = filepath.suffix.lower()
        if suffix in {".xlsx", ".xls"}:
            return cls._can_parse_excel(filepath)
        if suffix in {".csv", ""}:
            return cls._can_parse_csv(filepath)
        return False

    # -- CSV detection -------------------------------------------------------

    @classmethod
    def _can_parse_csv(cls, filepath: Path) -> bool:
        head = cls._read_head(filepath, 2048)
        if not head:
            return False
        lines = head.splitlines()
        if not lines:
            return False
        try:
            reader = csv.reader(io.StringIO(lines[0]))
            raw_headers = next(reader)
        except (StopIteration, csv.Error):
            return False
        norm = [_normalise_header(h) for h in raw_headers]
        if not (_find_col(norm, _CIDR_ALIASES) and _find_col(norm, _LABEL_ALIASES)):
            return False
        # Confirm at least one data row contains a valid CIDR in the CIDR column
        cidr_idx = norm.index(_find_col(norm, _CIDR_ALIASES))
        for line in lines[1:4]:
            try:
                row = next(csv.reader(io.StringIO(line)))
                if len(row) > cidr_idx and _is_valid_cidr(row[cidr_idx]):
                    return True
            except (StopIteration, csv.Error):
                continue
        return False

    # -- Excel detection -----------------------------------------------------

    @classmethod
    def _can_parse_excel(cls, filepath: Path) -> bool:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
            wb.close()
        except Exception:
            return False
        if not rows:
            return False
        norm = [_normalise_header(str(c or "")) for c in rows[0]]
        if not (_find_col(norm, _CIDR_ALIASES) and _find_col(norm, _LABEL_ALIASES)):
            return False
        cidr_idx = norm.index(_find_col(norm, _CIDR_ALIASES))
        for row in rows[1:]:
            val = str(row[cidr_idx] or "") if len(row) > cidr_idx else ""
            if _is_valid_cidr(val):
                return True
        return False

    # -- parse ---------------------------------------------------------------

    @classmethod
    def parse(cls, filepath: Path) -> ParseResult:
        result = ParseResult(source_file=str(filepath), parser_name=cls.name)
        suffix = filepath.suffix.lower()
        try:
            if suffix in {".xlsx", ".xls"}:
                cls._parse_excel(filepath, result)
            else:
                cls._parse_csv(filepath, result)
        except Exception as exc:
            result.errors.append(f"IPAM parse error: {exc}")
        if not result.errors:
            result.warnings.append(
                f"Imported {len(result.subnet_labels)} subnet definition(s) — "
                "no host records in an IPAM file."
            )
        return result

    @classmethod
    def _parse_csv(cls, filepath: Path, result: ParseResult) -> None:
        with open(filepath, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            raw_headers = reader.fieldnames or []
            norm_headers = [_normalise_header(h) for h in raw_headers]
            # Rebuild rows with normalised keys
            rows = [
                {_normalise_header(k): v for k, v in row.items()}
                for row in reader
            ]
        _parse_rows(rows, norm_headers, result)

    @classmethod
    def _parse_excel(cls, filepath: Path, result: ParseResult) -> None:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return
        norm_headers = [_normalise_header(str(c or "")) for c in all_rows[0]]
        rows = [
            {norm_headers[i]: str(cell or "") for i, cell in enumerate(row)}
            for row in all_rows[1:]
        ]
        _parse_rows(rows, norm_headers, result)
