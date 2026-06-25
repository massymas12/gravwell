"""Export callbacks — CSV, XLSX, graph PNG, and .gwexport project bundles."""
from __future__ import annotations
import base64
import csv
import io
from datetime import datetime

import dash
from dash import Input, Output, State, html, no_update, dcc
from flask import current_app
from flask_login import current_user
from gravwell.models.enrichment import resolve_vuln_name


# ── Query helpers ─────────────────────────────────────────────────────────────

def _fetch_data(db_path: str):
    """Return (host_rows, vuln_rows) as lists of dicts for export."""
    from gravwell.database import get_session
    from gravwell.models.orm import HostORM, VulnerabilityORM, CVERefORM, CVEEnrichmentORM

    host_rows = []
    vuln_rows = []

    with get_session(db_path) as session:
        hosts = session.query(HostORM).order_by(HostORM.ip).all()
        for h in hosts:
            open_ports = sorted(
                s.port for s in h.services if s.state == "open"
            )
            host_rows.append({
                "IP":           h.ip,
                "Hostnames":    "; ".join(h.hostnames),
                "OS Family":    h.os_family or "",
                "OS Name":      h.os_name or "",
                "Status":       h.status or "",
                "MAC":          h.mac or "",
                "MAC Vendor":   h.mac_vendor or "",
                "Open Ports":   "; ".join(str(p) for p in open_ports),
                "Max CVSS":     h.max_cvss or 0.0,
                "Critical":     h.vuln_count_critical or 0,
                "High":         h.vuln_count_high or 0,
                "Medium":       h.vuln_count_medium or 0,
                "Low":          h.vuln_count_low or 0,
                "Tags":         "; ".join(h.tags),
                "Notes":        (h.notes or "").replace("\n", " "),
            })

            for v in h.vulnerabilities:
                cve_ids = [r.cve_id for r in v.cve_refs]
                # KEV / EPSS: pick best signal across all CVEs on this vuln
                in_kev = False
                kev_date = ""
                epss_score = ""
                epss_pct = ""
                if cve_ids:
                    enrichments = (
                        session.query(CVEEnrichmentORM)
                        .filter(CVEEnrichmentORM.cve_id.in_(
                            [c.upper() for c in cve_ids]
                        ))
                        .all()
                    )
                    enrich_map = {}
                    for e in enrichments:
                        enrich_map[e.cve_id.upper()] = e
                        if e.in_kev:
                            in_kev = True
                            kev_date = e.kev_date_added or ""
                        if e.epss_score is not None:
                            cur = float(e.epss_score)
                            if epss_score == "" or cur > float(epss_score):
                                epss_score = round(cur, 4)
                                epss_pct = round(float(e.epss_percentile or 0), 4)

                vuln_rows.append({
                    "IP":          h.ip,
                    "Hostname":    h.hostnames[0] if h.hostnames else "",
                    "OS Family":   h.os_family or "",
                    "Port":        v.port or "",
                    "Plugin ID":   v.plugin_id or "",
                    "CVE IDs":     "; ".join(cve_ids),
                    "Name":        resolve_vuln_name(v.name or "", cve_ids, enrich_map, 256),
                    "Severity":    v.severity or "",
                    "CVSS":        v.cvss_score or "",
                    "In KEV":      "Yes" if in_kev else "",
                    "KEV Date":    kev_date,
                    "EPSS Score":  epss_score,
                    "EPSS %ile":   epss_pct,
                    "Description": (v.description or "")[:1000],
                    "Solution":    (v.solution or "")[:500],
                })

    return host_rows, vuln_rows


def _to_csv(host_rows: list[dict], vuln_rows: list[dict]) -> str:
    """Single CSV file with a Hosts section then a Vulnerabilities section."""
    buf = io.StringIO()
    w = csv.writer(buf)

    if host_rows:
        w.writerow(["# HOSTS"])
        w.writerow(list(host_rows[0].keys()))
        for r in host_rows:
            w.writerow(list(r.values()))
        w.writerow([])

    w.writerow(["# VULNERABILITIES"])
    if vuln_rows:
        w.writerow(list(vuln_rows[0].keys()))
        for r in vuln_rows:
            w.writerow(list(r.values()))

    return buf.getvalue()


def _to_xlsx(host_rows: list[dict], vuln_rows: list[dict]) -> bytes:
    """Two-sheet XLSX workbook: Hosts + Vulnerabilities."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT = Font(bold=True, color="5DADE2", size=10)
    SEV_COLOURS = {
        "critical": "E74C3C",
        "high":     "E67E22",
        "medium":   "F1C40F",
        "low":      "27AE60",
        "info":     "5DADE2",
    }

    def _write_sheet(ws, rows: list[dict], title: str):
        ws.title = title
        if not rows:
            ws.append(["No data"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append(list(row.values()))

        # Auto-width (capped at 60)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # ── Hosts sheet ───────────────────────────────────────────────────────
    ws_hosts = wb.active
    _write_sheet(ws_hosts, host_rows, "Hosts")

    # ── Vulns sheet ───────────────────────────────────────────────────────
    ws_vulns = wb.create_sheet("Vulnerabilities")
    _write_sheet(ws_vulns, vuln_rows, "Vulnerabilities")

    # Colour-code severity column
    try:
        sev_col = list(vuln_rows[0].keys()).index("Severity") + 1 if vuln_rows else None
        if sev_col:
            for row in ws_vulns.iter_rows(min_row=2, min_col=sev_col, max_col=sev_col):
                for cell in row:
                    sev = str(cell.value or "").lower()
                    colour = SEV_COLOURS.get(sev)
                    if colour:
                        cell.fill = PatternFill("solid", fgColor=colour)
                        cell.font = Font(bold=True, color="FFFFFF", size=9)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Callbacks ─────────────────────────────────────────────────────────────────

def register(app: dash.Dash) -> None:

    @app.callback(
        Output("export-download", "data"),
        Output("hamburger-menu", "style", allow_duplicate=True),
        Output("hamburger-backdrop", "style", allow_duplicate=True),
        Input("export-csv-menu-item", "n_clicks"),
        Input("export-xlsx-menu-item", "n_clicks"),
        prevent_initial_call=True,
    )
    def export_data(csv_clicks, xlsx_clicks):
        from dash import ctx
        triggered = ctx.triggered_id
        if not triggered:
            return no_update, no_update, no_update

        closed = {"display": "none"}

        if not current_user.is_authenticated or not current_user.can("export"):
            return no_update, closed, closed

        db_path = current_app.config.get("GRAVWELL_DB_PATH", "")
        host_rows, vuln_rows = _fetch_data(db_path)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if triggered == "export-csv-menu-item":
            content = _to_csv(host_rows, vuln_rows)
            return (
                dcc.send_string(content, f"gravwell_{stamp}.csv",
                                type="text/csv"),
                closed, closed,
            )
        else:
            content = _to_xlsx(host_rows, vuln_rows)
            return (
                dcc.send_bytes(content, f"gravwell_{stamp}.xlsx"),
                closed, closed,
            )

    # PNG export is handled entirely client-side; this callback closes the menu
    @app.callback(
        Output("export-png-dummy", "data"),
        Output("hamburger-menu", "style", allow_duplicate=True),
        Output("hamburger-backdrop", "style", allow_duplicate=True),
        Input("export-png-menu-item", "n_clicks"),
        prevent_initial_call=True,
    )
    def trigger_png_export(n_clicks):
        if not n_clicks:
            return no_update, no_update, no_update
        # The actual download is done by a clientside callback watching this store
        return (
            {"_t": n_clicks},
            {"display": "none"},
            {"display": "none"},
        )

    # ── Map export (.drawio) ─────────────────────────────────────────────────

    @app.callback(
        Output("drawio-export-download", "data"),
        Output("hamburger-menu", "style", allow_duplicate=True),
        Output("hamburger-backdrop", "style", allow_duplicate=True),
        Output("drawio-export-status", "children"),
        Input("export-drawio-menu-item", "n_clicks"),
        prevent_initial_call=True,
    )
    def trigger_drawio_export(n_clicks):
        if not n_clicks:
            return no_update, no_update, no_update, no_update
        if not current_user.is_authenticated or not current_user.can("export"):
            return no_update, no_update, no_update, no_update
        from gravwell.drawio import export_drawio
        db_path = current_app.config["GRAVWELL_DB_PATH"]
        try:
            xml = export_drawio(db_path)
        except Exception as e:
            return no_update, {"display": "none"}, {"display": "none"}, ""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return (
            dcc.send_string(xml, f"network_map_{stamp}.drawio"),
            {"display": "none"},
            {"display": "none"},
            "",
        )

    # ── Project export (.gwexport) ────────────────────────────────────────────

    @app.callback(
        Output("export-project-modal-overlay", "style"),
        Output("hamburger-menu", "style", allow_duplicate=True),
        Output("hamburger-backdrop", "style", allow_duplicate=True),
        Input("export-project-menu-item", "n_clicks"),
        prevent_initial_call=True,
    )
    def open_export_project_modal(n_clicks):
        if not n_clicks:
            return no_update, no_update, no_update
        closed = {"display": "none"}
        return {"display": "flex"}, closed, closed

    @app.callback(
        Output("export-project-modal-overlay", "style", allow_duplicate=True),
        Output("export-project-passphrase", "value"),
        Output("export-project-passphrase-confirm", "value"),
        Output("export-project-error", "children"),
        Input("export-project-modal-close", "n_clicks"),
        Input("cancel-export-project-btn", "n_clicks"),
        prevent_initial_call=True,
    )
    def close_export_project_modal(_close, _cancel):
        return {"display": "none"}, "", "", ""

    @app.callback(
        Output("project-export-download", "data"),
        Output("export-project-modal-overlay", "style", allow_duplicate=True),
        Output("export-project-passphrase", "value", allow_duplicate=True),
        Output("export-project-passphrase-confirm", "value", allow_duplicate=True),
        Output("export-project-error", "children", allow_duplicate=True),
        Input("confirm-export-project-btn", "n_clicks"),
        State("export-project-passphrase", "value"),
        State("export-project-passphrase-confirm", "value"),
        prevent_initial_call=True,
    )
    def do_project_export(n_clicks, passphrase, confirm):
        if not n_clicks:
            return no_update, no_update, no_update, no_update, no_update
        if not current_user.is_authenticated or not current_user.can("export"):
            return no_update, no_update, no_update, no_update, "Not authorized."
        passphrase = (passphrase or "").strip()
        confirm    = (confirm    or "").strip()
        if not passphrase:
            return no_update, no_update, no_update, no_update, "Passphrase cannot be empty."
        if passphrase != confirm:
            return no_update, no_update, no_update, no_update, "Passphrases do not match."
        from gravwell.export import export_project, encrypt_bundle
        db_path = current_app.config["GRAVWELL_DB_PATH"]
        try:
            blob = encrypt_bundle(export_project(db_path), passphrase)
        except Exception as e:
            return no_update, no_update, no_update, no_update, f"Export failed: {e}"
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return (
            dcc.send_bytes(blob, f"gravwell_export_{stamp}.gwexport"),
            {"display": "none"},
            "", "",
            "",
        )

    # ── Project import (.gwexport) ────────────────────────────────────────────

    @app.callback(
        Output("import-project-modal-overlay", "style", allow_duplicate=True),
        Output("import-project-passphrase", "value"),
        Output("import-project-error", "children"),
        Input("import-project-modal-close", "n_clicks"),
        Input("cancel-import-project-btn", "n_clicks"),
        prevent_initial_call=True,
    )
    def close_import_project_modal(_close, _cancel):
        return {"display": "none"}, "", ""

    @app.callback(
        Output("import-project-filename-label", "children"),
        Input("gwexport-pending-store", "data"),
        prevent_initial_call=True,
    )
    def update_import_filename_label(pending):
        if not pending:
            return ""
        filename = pending.get("filename", "export file")
        return f"File: {filename}"

    @app.callback(
        Output("import-project-modal-overlay", "style", allow_duplicate=True),
        Output("import-project-passphrase", "value", allow_duplicate=True),
        Output("import-project-error", "children", allow_duplicate=True),
        Output("upload-status", "children", allow_duplicate=True),
        Output("scan-file-list", "children", allow_duplicate=True),
        Output("data-refresh-trigger", "data", allow_duplicate=True),
        Input("confirm-import-project-btn", "n_clicks"),
        State("import-project-passphrase", "value"),
        State("gwexport-pending-store", "data"),
        State("data-refresh-trigger", "data"),
        prevent_initial_call=True,
    )
    def do_project_import(n_clicks, passphrase, pending, refresh_counter):
        if not n_clicks or not pending:
            return no_update, no_update, no_update, no_update, no_update, no_update
        if not current_user.is_authenticated or not current_user.can("import"):
            return no_update, no_update, "Not authorized.", no_update, no_update, no_update
        passphrase = (passphrase or "").strip()
        if not passphrase:
            return no_update, no_update, "Passphrase cannot be empty.", no_update, no_update, no_update
        from gravwell.export import decrypt_bundle, import_project
        from gravwell.ui.callbacks.import_callbacks import _build_scan_file_list
        db_path  = current_app.config["GRAVWELL_DB_PATH"]
        filename = pending.get("filename", "export")
        try:
            raw        = base64.b64decode(pending["contents_b64"])
            json_bytes = decrypt_bundle(raw, passphrase)
            h_count, v_count = import_project(json_bytes, db_path)
        except ValueError as e:
            return no_update, no_update, str(e), no_update, no_update, no_update
        except Exception as e:
            return no_update, no_update, f"Import failed: {e}", no_update, no_update, no_update
        status = html.Div(
            f"Imported {filename}: {h_count} hosts, {v_count} vulns",
            style={"color": "#27AE60", "fontSize": "12px"},
        )
        return (
            {"display": "none"},
            "",
            "",
            status,
            _build_scan_file_list(db_path),
            (refresh_counter or 0) + 1,
        )
